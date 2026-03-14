from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
import json
from .Recommand import recommendCommunication
from .Recommand import llm_summary
from .Recommand import get_latest_entries
@login_required
def home_view(request):
    """Main recommendation page"""
    return render(request, 'recommendations/home.html')


@login_required
@require_http_methods(["POST"])
def get_recommendation_ajax(request):
    """AJAX endpoint for getting recommendations"""
    try:
        data = json.loads(request.body)
        book_name = data.get('book_name', '')
        author_name = data.get('author_name', '')
        if not book_name:
            return JsonResponse({
                'success': False,
                'error': 'Please enter a book name'
            })

        recommendations = recommendCommunication(book_name,author_name, n_recommendations=10)

        if not recommendations:
            return JsonResponse({
                'success': False,
                'error': 'No recommendations found'
            })

        return JsonResponse({
            'success': True,
            'book_name': book_name,
            'recommendations': recommendations
        })


    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
from django.core.mail import send_mail
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
import json

@login_required
def email_recommendations(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            book_name = data.get('book_name', '')
            recommendations = data.get('recommendations', [])
            user_email = request.user.email

            if not recommendations:
                return JsonResponse({'success': False, 'error': 'No recommendations found'})

            message = f"Here are your book recommendations based on '{book_name}':\n\n"
            message += "\n".join([f"{i+1}. {book}" for i, book in enumerate(recommendations)])

            send_mail(
                subject=f"Book Recommendations for '{book_name}'",
                message=message,
                from_email='yourapp@example.com',
                recipient_list=[user_email],
                fail_silently=False,
            )

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
@require_http_methods(["POST"])
def email_summary(request):
    data = json.loads(request.body)

    title = data["title"]
    author = data.get("author", "")
    summary = data["summary"]

    send_mail(
        subject=f"Summary: {title}",
        message=f"{title}\nby {author}\n\n{summary}",
        from_email="yourapp@example.com",
        recipient_list=[request.user.email],
    )

    return JsonResponse({"success": True})


from django.http import JsonResponse
from .models import BookSummary
from .Recommand import llm_summary
import json

def get_summary(request):
    data = json.loads(request.body)
    title = data["title"]
    author = data.get("author", "")

    #Try cache
    cached = BookSummary.objects.filter(
        title__iexact=title,
        author__iexact=author
    ).first()

    if cached:
        return JsonResponse({"summary": cached.summary, "cached": True})

    # Generate with Groq
    prompt = f"""
    Write a concise, spoiler-free summary of the book:
    Title: {title}
    Author: {author}
    Length: 3 sentences
    Tone: engaging and neutral
    """

    summary = llm_summary(prompt)

    if not summary:
        return JsonResponse({"error": "Failed to retrieve summary"}, status=500)

    # save to cache
    BookSummary.objects.create(
        title=title,
        author=author,
        summary=summary
    )

    return JsonResponse({"summary": summary, "cached": False})


from openpyxl import Workbook
from openpyxl.styles import Font
import tempfile
import os


def create_books_excel(books, authors, quotes):
    wb = Workbook()

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(title=name)
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in rows:
            ws.append(row)

        ws.auto_filter.ref = ws.dimensions

    wb.remove(wb.active)

    add_sheet(
        "Books",
        ["ID", "Title", "Author", "Notes", "Number of registered quotes"],
        books
    )

    add_sheet(
        "Authors",
        ["ID", "Name", "Notes", "Number of books registered", "Number of quotes registered"],
        authors
    )

    add_sheet(
        "Quotes",
        [
            "ID", "Book", "Author",
            "Quote","Notes",
            "Start Page", "Start Row",
            "End Page", "End Row",
            "Keywords"
        ],
        quotes
    )

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    temp_file.close()

    return temp_file.name

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMessage

@login_required
def send_notif(request):
    books = get_latest_entries("books")
    authors = get_latest_entries("authors")
    quotes = get_latest_entries("quotes")

    excel_path = create_books_excel(books, authors, quotes)

    email = EmailMessage(
        subject="📚 Latest entries from your Book Database",
        body=(
            "Hello!\n\n"
            "Attached you will find an Excel file containing:\n"
            "- Latest books\n"
            "- Authors\n"
            "- Quotes\n\n"
            "Enjoy reading!"
        ),
        to=[request.user.email],
    )

    email.attach_file(excel_path)
    email.send()
    import os
    os.remove(excel_path)
    return JsonResponse({"success": True})

